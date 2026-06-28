#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_data.py — HRBP Workforce & Recruitment Command Centre
================================================================================
One-time, monthly build script. Reads the three sanitised HR workbooks, parses
them section-by-section (they are human-formatted reports, NOT clean tables),
applies the HRBP reconciliation mapping, masks all PII at bake time, computes
KPIs + a data-quality report, and emits  js/data.js  as:

        window.DASHBOARD_DATA = { ... };

The dashboard then loads that JS global directly — so it works on GitHub Pages
(https, sub-path) AND when index.html is opened from disk (file://), with zero
network calls at runtime.

USAGE
-----
    pip install openpyxl
    python build/build_data.py            # reads xlsx from repo root or ./data
    python build/build_data.py --src .    # explicit source dir

PRIVACY
-------
Names / employee codes / candidate names / supervisor refs are masked to stable,
non-reversible pseudonyms (Employee Group N, Candidate A, Manager NN). After
generation the emitted data.js is scanned and the build FAILS LOUDLY if any
PII-shaped string (a source name or a 6–8 digit emp-code) leaks. Safe to publish.
================================================================================
"""

import argparse
import collections
import datetime as _dt
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl is required.  Run:  pip install openpyxl")

# ------------------------------------------------------------------------------
# 0. CONFIG
# ------------------------------------------------------------------------------
FILE_REVIEW = "Monthly HRBP Review.xlsx"
FILE_BUDGET = "Budgted numbers Corp and SSC.xlsx"
FILE_TRACKER = "Recruitment Tracker.xlsx"

DEFAULT_TAT_DAYS = 45          # tracker's TAT field is scrambled -> assumed default
NOT_AVAILABLE = "Not Available"

# ------------------------------------------------------------------------------
# HRBP reconciliation — DATA-DRIVEN and NAME-FREE.
#
# IMPORTANT (privacy): this file is published in a PUBLIC repo, so it must NOT
# contain any real employee / HRBP names. Portfolios are discovered entirely at
# run time from the workbooks the user supplies — nothing about the org is baked
# into the source. The two real-world wrinkles are handled by generic rules, not
# by a hard-coded name list:
#
#   1. Spelling variants of one person (e.g. a short form and a longer form of
#      the same name) are merged with a generic PREFIX rule: if one portfolio
#      name is a case-insensitive prefix of another (shared stem >= 4 chars),
#      they are treated as the same person and the SHORTER form is the display.
#
#   2. A non-person "pool" detail sheet (trainee / OJT / SSC / intern pool) is
#      detected from generic keyword TOKENS in the sheet name and folded into the
#      largest person portfolio, so its head-count rolls up correctly instead of
#      appearing as a phantom HRBP. (Override-able via the optional, git-ignored
#      build/portfolios.local.json — see that file's .example — for orgs where
#      the pool belongs elsewhere.)
#
# The result: no names in the repo, and the same reconciliation on real data.

# Keyword tokens that mark a budget detail sheet as a non-person "pool" sheet.
POOL_TOKENS = ("ojt", "ssc", "trainee", "intern", "apprentice", " pool", "get pool")
# Minimum shared-stem length for the generic prefix-merge of name variants.
_STEM_MIN = 4

def canon(name):
    """Normalise a name for matching: trim + collapse internal whitespace.
    (No name mapping — variant merging is done generically by build_resolver.)"""
    if name is None:
        return ""
    return re.sub(r"\s+", " ", str(name)).strip()

def is_pool_name(name):
    """True if a budget sheet name looks like a non-person trainee/OJT pool."""
    lw = " " + canon(name).lower() + " "
    return any(tok in lw for tok in POOL_TOKENS)

def _prefix_match(a, b):
    """True if a and b share a stem: one is a case-insensitive prefix of the
    other and the shorter is at least _STEM_MIN chars."""
    if not a or not b:
        return False
    short = a if len(a) <= len(b) else b
    return len(short) >= _STEM_MIN and (a.startswith(b) or b.startswith(a))

def build_resolver(names):
    """Group candidate person names so that prefix-variants of one name collapse
    to a single canonical key (the SHORTEST form becomes the display). Returns
    (resolve, displays): resolve(name) -> portfolio key or None; displays maps
    key -> display name. Entirely data-driven — no hard-coded names."""
    forms = {}
    for n in names:
        c = canon(n)
        if c:
            forms.setdefault(c.lower(), c)
    stems, canon_lower = [], {}
    for lw in sorted(forms, key=len):                 # shortest first
        hit = next((s for s in stems if _prefix_match(s, lw)), None)
        if hit:
            canon_lower[lw] = hit
        else:
            stems.append(lw)
            canon_lower[lw] = lw
    key_of = {s: slug(forms[s]) for s in stems}
    displays = {key_of[s]: forms[s] for s in stems}

    def resolve(name):
        c = canon(name).lower()
        if not c:
            return None
        if c in canon_lower:
            return key_of[canon_lower[c]]
        hit = next((s for s in stems if _prefix_match(s, c)), None)
        return key_of[hit] if hit else None
    return resolve, displays

def load_local_overrides(src):
    """Optional, git-ignored per-deployment config so an org can pin specific
    name merges / pool-fold targets without putting names in the repo. Looks for
    build/portfolios.local.json next to the script or in <src>. Shape:
        {"fold": {"<pool sheet name>": "<target display>"},
         "merge": {"<variant>": "<canonical display>"}}
    Returns (fold, merge) lower-cased dicts (empty if the file is absent)."""
    fold, merge = {}, {}
    here = os.path.dirname(os.path.abspath(__file__))
    for cand in (os.path.join(here, "portfolios.local.json"),
                 os.path.join(src or ".", "portfolios.local.json")):
        if os.path.exists(cand):
            try:
                with open(cand, encoding="utf-8") as fh:
                    cfg = json.load(fh)
                fold = {str(k).strip().lower(): v for k, v in (cfg.get("fold") or {}).items()}
                merge = {str(k).strip().lower(): v for k, v in (cfg.get("merge") or {}).items()}
                print("  loaded local portfolio overrides:", cand)
            except Exception as e:                       # noqa: BLE001
                warn("could not read %s: %s" % (cand, e))
            break
    return fold, merge

# Sheets to detect-but-ignore for record parsing (pivots / scratch / empty).
IGNORE_SHEET_HINTS = [
    "sheet1", "sheet2", "sheet3", "sheet4", "sheet5", "sheet6", "sheet7",
    "function wise", "birthday celebration", "hrbp wise summary",
    "overall summary", "finance", "recruitment tracker finance",
]

WARNINGS = []
def warn(msg):
    WARNINGS.append(msg)
    print("  [warn]", msg)


# ------------------------------------------------------------------------------
# small helpers
# ------------------------------------------------------------------------------
def norm(s):
    """Trim + collapse whitespace/newlines of a header/label."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\n", " ")).strip()

def is_blank(v):
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s in ("-", "NA", "N/A", "na", "n/a", "#REF!")

def num(v, default=None):
    """Best-effort numeric extraction (handles '251.62 days', '6.19', 0.026...)."""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace(",", ""))
    return float(m.group()) if m else default

def pct(frac):
    """0.026 -> 2.6 ; passthrough if already a percentage-ish number."""
    if frac is None:
        return None
    f = float(frac)
    return round(f * 100, 1) if f <= 1.0 else round(f, 1)

def find_sheet(wb, *hints):
    """Case-insensitive, trimmed, contains-based sheet matcher."""
    lc = {name: name.strip().lower() for name in wb.sheetnames}
    for h in hints:
        h = h.strip().lower()
        for name, low in lc.items():
            if low == h:
                return name
    for h in hints:
        h = h.strip().lower()
        for name, low in lc.items():
            if h in low or low in h:
                return name
    return None


# ==============================================================================
# 1. PRIVACY / MASKING LAYER  (applied at bake time AND mirrored in upload.js)
# ==============================================================================
class Masker:
    """Stable, non-reversible pseudonymisation. Same source value -> same label
    within a build, but labels carry no real-world meaning."""
    def __init__(self):
        self._emp = {}
        self._cand = {}
        self._mgr = {}

    def employee(self, raw):                       # Emp. Name -> "Employee Group N"
        if is_blank(raw):
            return None
        key = str(raw).strip().lower()
        if key not in self._emp:
            self._emp[key] = "Employee Group %d" % (len(self._emp) + 1)
        return self._emp[key]

    def candidate(self, raw):                      # Candidate Name -> "Candidate A.."
        if is_blank(raw):
            return None
        key = str(raw).strip().lower()
        if key not in self._cand:
            n = len(self._cand)
            # A..Z, then AA, AB ...
            label = ""
            x = n
            while True:
                label = chr(ord("A") + x % 26) + label
                x = x // 26 - 1
                if x < 0:
                    break
            self._cand[key] = "Candidate %s" % label
        return self._cand[key]

    def manager(self, raw):                        # Reports to / Supervisor -> "Manager NN"
        if is_blank(raw):
            return None
        key = str(raw).strip().lower()
        if key not in self._mgr:
            self._mgr[key] = "Manager %02d" % (len(self._mgr) + 1)
        return self._mgr[key]

    def role(self, idx):                            # masked record id for tracker rows
        return "Role %04d" % idx

    def source_name_values(self):
        """All raw name strings seen — used by the output guard."""
        out = set()
        for d in (self._emp, self._cand, self._mgr):
            out.update(d.keys())
        return out

MASK = Masker()


# ==============================================================================
# 2. MONTHLY REVIEW  ('Summary' sheet) — the PRIMARY source.
#    Parsed SECTION BY SECTION using the label in col B/C as the anchor.
# ==============================================================================
def parse_review(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = find_sheet(wb, "Summary")
    ws = wb[sn]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    def cell(r, c):
        if 0 <= r < len(rows) and 0 <= c < len(rows[r]):
            return rows[r][c]
        return None

    # locate a section header (text in col B, index 1) -> returns its row index
    def section_row(label):
        for i, r in enumerate(rows):
            if norm(cell(i, 1)).lower() == label.lower():
                return i
        return None

    def label_in_row(i):
        # section row-labels live in col C (index 2)
        return norm(cell(i, 2))

    # Discover the HRBP names from a section's HEADER ROW rather than hard-coding
    # them. `span` is the columns each HRBP occupies (1 normally; 3 in the
    # Recruitment Overview where each HRBP splits into RPO/Consultant/Other).
    def read_hrbp_header(header_row, span=1, start_col=3):
        pairs, c = [], start_col
        while c < 40:
            v = norm(cell(header_row, c))
            if not v or v.lower() in ("total", "grand total"):
                break
            pairs.append((v, c))
            c += span
        return pairs

    out = {"hrbps": {}, "engagement": [], "initiatives": [], "benchmark": None}

    # ---- Headcount Overview (authoritative HRBP name list) -------------------
    hc_start = section_row("Headcount Overview")
    hc_pairs = read_hrbp_header(hc_start + 1) if hc_start is not None else []
    hrbp_names = [p[0] for p in hc_pairs]        # HRBP labels read from the sheet header
    hc_cols = {name: col for name, col in hc_pairs}
    out["hrbpNames"] = hrbp_names
    for h in hrbp_names:
        out["hrbps"][h] = {}
    hc_fields = {
        "Total Position 2026": "budget",
        "Active Employees": "active",
        "Joining's in June": "joiningsJune",
        "Exits in June": "exitsJune",
        "Joining's YTD": "joiningsYTD",
        "Exits YTD": "exitsYTD",
        "Current Positions": "wipCurrent",
        "Future Positions": "future",
        "On Hold": "onHold",
        "Delimit/Redundant": "delimit",
        "Attrition %": "attrition",
        "Insights on Attrition": "attritionInsight",
    }
    if hc_start is not None:
        for i in range(hc_start + 1, hc_start + 16):
            lbl = label_in_row(i)
            if not lbl:
                continue
            for key, field in hc_fields.items():
                if key.lower() in lbl.lower():
                    for h, c in hc_cols.items():
                        v = cell(i, c)
                        if field == "attrition":
                            out["hrbps"][h][field] = pct(num(v))
                        elif field == "attritionInsight":
                            out["hrbps"][h][field] = (norm(v) if not is_blank(v)
                                                      else NOT_AVAILABLE)
                        else:
                            out["hrbps"][h][field] = (int(num(v)) if num(v) is not None
                                                      else None)
                    break

    # ---- Recruitment Overview (RPO / Consultant / Other per HRBP) ------------
    # header row has each HRBP spanning 3 sub-cols. We locate the HRBP header row
    # then read its 3 columns. Stages map onto the funnel.
    rec_start = section_row("Recruitment Overview")
    # EXACT-label map (anchored on col C). Note "Offered" is a substring of
    # "To Be Offered", so we match on normalised equality, NOT contains.
    rec_field = {"wip": "wip", "offered": "offered", "to be offered": "toBeOffered",
                 "joined 2026": "joined", "offer declined": "offerDeclined"}
    # each HRBP spans 3 sub-columns (RPO / Consultant / Other) — discover starts.
    rec_cols = ({name: col for name, col in read_hrbp_header(rec_start + 1, span=3)}
                if rec_start is not None else {})
    if rec_start is not None:
        for h in hrbp_names:
            out["hrbps"].setdefault(h, {})["recruitment"] = {}
        for i in range(rec_start + 1, rec_start + 9):
            lbl = label_in_row(i).lower().strip()
            field = rec_field.get(lbl)
            if field is None:
                continue
            for h, c0 in rec_cols.items():
                rpo = num(cell(i, c0)) or 0
                cons = num(cell(i, c0 + 1)) or 0
                other = num(cell(i, c0 + 2)) or 0
                out["hrbps"][h]["recruitment"][field] = {
                    "rpo": int(rpo), "consultant": int(cons), "other": int(other),
                    "total": int(rpo + cons + other),
                }

    # ---- Aging Overview ------------------------------------------------------
    ag_start = section_row("Aging Overview")
    ag_cols = ({name: col for name, col in read_hrbp_header(ag_start + 1)}
               if ag_start is not None else {})
    ag_rows = {"WIP": "wip", "0-30": "b0_30", "30-60": "b30_60",
               "60-90": "b60_90", "90 +": "b90"}
    if ag_start is not None:
        for h in hrbp_names:
            out["hrbps"].setdefault(h, {})["aging"] = {}
        for i in range(ag_start + 1, ag_start + 9):
            lbl = label_in_row(i)
            if not lbl:
                continue
            if lbl.lower().startswith("critical"):
                for h, c in ag_cols.items():
                    txt = cell(i, c)
                    out["hrbps"][h]["criticalCases"] = (norm(txt).split("  ")
                                                        if not is_blank(txt) else [])
                    if not is_blank(txt):
                        # split numbered list "1. .. 2. .. 3. .."
                        parts = re.split(r"\s*\d+\.\s*", str(txt))
                        out["hrbps"][h]["criticalCases"] = [norm(p) for p in parts if norm(p)]
                continue
            for ak, field in ag_rows.items():
                if ak.lower() == lbl.lower() or ak.replace(" ", "") == lbl.replace(" ", ""):
                    for h, c in ag_cols.items():
                        out["hrbps"][h]["aging"][field] = int(num(cell(i, c)) or 0)
                    break

    # ---- HR Initiatives Calendar (ragged grid) ------------------------------
    init_start = section_row("HR Initiatives Calendar")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sept", "Oct", "Nov", "Dec"]
    if init_start is not None:
        # header months sit at init_start+1, cols 3..14
        end = section_row("Training & Development") or (init_start + 40)
        current_hrbp = None
        for i in range(init_start + 2, end):
            row_label = norm(cell(i, 2))
            if row_label:
                # detect an HRBP block label (may be a bare name or "<name> for SSC", etc.)
                matched = None
                for h in hrbp_names:
                    if h.lower() in row_label.lower():
                        matched = h
                        break
                if matched:
                    current_hrbp = matched
                # if not a known HRBP, the label is itself an event chip (col C event)
            for mi, m in enumerate(months):
                v = cell(i, 3 + mi)
                if not is_blank(v):
                    txt = norm(v)
                    if txt and not txt.lower().startswith("dummy business note"):
                        out["initiatives"].append({
                            "hrbp": current_hrbp or NOT_AVAILABLE,
                            "month": m, "event": txt,
                            "category": classify_event(txt),
                        })

    # ---- Training & Development ----------------------------------------------
    tr_start = section_row("Training & Development")
    if tr_start is not None:
        # header at tr_start+1; rows tr_start+2 .. until 'Org Charts'
        end = section_row("Org Charts") or (tr_start + 10)
        # Row labels in col C may be scrambled; capture every data row that has a
        # numeric Total Headcount, then attach by the HRBP name in the label.
        train_rows = []
        for i in range(tr_start + 2, end):
            total = num(cell(i, 3))
            label = norm(cell(i, 2))
            if total is None:
                continue
            days = num(cell(i, 9))
            train_rows.append({
                "rowLabel": label, "totalHeadcount": int(total),
                "trainingDays": round(days, 2) if days is not None else None,
                "upcoming": norm(cell(i, 10)) or NOT_AVAILABLE,
            })
        # Attach by matching the HRBP name embedded in the row label where possible.
        for h in hrbp_names:
            for tr in train_rows:
                if h.lower() in tr["rowLabel"].lower():
                    out["hrbps"][h]["training"] = tr
                    break
        # Fallback: if the first HRBP's row label was scrambled, take the first row.
        if hrbp_names and "training" not in out["hrbps"].get(hrbp_names[0], {}) and train_rows:
            out["hrbps"].setdefault(hrbp_names[0], {})["training"] = train_rows[0]

    # ---- Org Charts ----------------------------------------------------------
    oc_start = section_row("Org Charts")
    if oc_start is not None:
        for i in range(oc_start + 1, oc_start + 8):
            lbl = norm(cell(i, 2))
            for h in hrbp_names:
                if lbl.lower() == h.lower():
                    out["hrbps"][h]["orgChart"] = norm(cell(i, 3)) or NOT_AVAILABLE

    # ---- PMS -----------------------------------------------------------------
    pms_start = section_row("PMS")
    if pms_start is not None:
        for i in range(pms_start + 1, pms_start + 8):
            lbl = norm(cell(i, 2))
            for h in hrbp_names:
                if lbl.lower() == h.lower():
                    out["hrbps"][h]["pms"] = {
                        "goalSetting": pct(num(cell(i, 3))),
                        "midYear": norm(cell(i, 4)) or NOT_AVAILABLE,
                        "endYear": norm(cell(i, 5)) or NOT_AVAILABLE,
                    }

    # ---- Speak-Up ------------------------------------------------------------
    su_start = None
    for i, r in enumerate(rows):
        if "speak" in norm(cell(i, 1)).lower() and "up" in norm(cell(i, 1)).lower():
            su_start = i
            break
    if su_start is not None:
        # header row su_start+1 lists the HRBPs (order may differ from Headcount).
        su_cols = {}
        for name, col in read_hrbp_header(su_start + 1):
            for h in hrbp_names:
                if name.lower() == h.lower():
                    su_cols[h] = col
        milestones = []
        for i in range(su_start + 2, su_start + 12):
            lbl = norm(cell(i, 2))
            if not lbl:
                continue
            if "listening" in lbl.lower() or "speak up+" in lbl.lower() or "scores" in lbl.lower():
                break
            ms = {"milestone": lbl, "status": {}}
            for h, c in su_cols.items():
                st = norm(cell(i, c))
                ms["status"][h] = st if st else "Pending"
            milestones.append(ms)
        out["speakUpMilestones"] = milestones

    # ---- Speak-Up / Listening scores: company benchmark + per-department -----
    # Anchor on the table header ("Department" + "...Survey..."). The company
    # benchmark (e.g. "AMNS India") is the scored row BEFORE the first HRBP label.
    el_hdr = None
    for i in range(len(rows)):
        if norm(cell(i, 2)).lower() == "department" and "survey" in norm(cell(i, 3)).lower():
            el_hdr = i
            break
    if el_hdr is not None:
        current_hrbp = NOT_AVAILABLE
        seen_hrbp = False
        for i in range(el_hdr + 1, len(rows)):
            hrbp_cell = norm(cell(i, 1))
            dept = norm(cell(i, 2))
            score = num(cell(i, 3))
            if not hrbp_cell and is_blank(dept) and score is None:
                if seen_hrbp:
                    break          # blank row after the data -> section end
                continue
            low_dept = dept.lower()
            if "top talent" in low_dept or "readiness" in low_dept or "hipo" in low_dept:
                break              # next section
            if hrbp_cell:
                current_hrbp = hrbp_cell
                seen_hrbp = True
            if not seen_hrbp and score is not None:
                out["benchmark"] = score      # company-level benchmark row
                continue
            if is_blank(dept) or len(dept) > 60:
                continue
            out["engagement"].append({
                "hrbp": current_hrbp,
                "department": dept,
                "score": score,
                "scoreLabel": (str(round(score, 2)) if score is not None else NOT_AVAILABLE),
            })

    return out


EVENT_CATEGORIES = [
    ("Compliance", ["compliance", "safety", "ethics", "posh", "audit", "vishwakarma", "loto",
                    "work at height", "transgender", "data hygiene", "data clean"]),
    ("PMS", ["goal setting", "goal settings", "mid-year", "mid year", "end - year", "end-year",
             "annual review", "kpi", "apr ", "performance", "appraisal", "my commitment", "commitments"]),
    ("Capability", ["hr xcel", "synergy meet", "training", "learning", "knowledge", "workshop",
                    "samvaad", "academy", "mentor", "succeed", "seed", "lead", "leap", "induction"]),
    ("Communication", ["sampark", "time out with hod", "coffee with", "interaction", "town hall",
                       "feedback", "connect", "praise points", "report sharing"]),
    ("Engagement", ["speakup", "speak up", "speak-up", "survey", "recognition", "reward", "r&r",
                    "felicitation", "promotion", "celebration", "festival", "diwali", "ganesh",
                    "navratri", "holi", "christmas", "bday", "birthday", "farewell", "retirement",
                    "sports", "games", "competition", "carrom", "chess", "drawing", "pot lunch",
                    "health", "yoga", "blood donation", "wellbeing", "environment", "plantation",
                    "csr", "voluntary", "new year", "republic day", "independence day", "sankranti",
                    "women's day", "quality month"]),
]
def classify_event(text):
    t = text.lower()
    for cat, kws in EVENT_CATEGORIES:
        if any(k in t for k in kws):
            return cat
    return "Culture"


# ==============================================================================
# 3. BUDGET WORKBOOK — record-level headcount/budget (detail sheets).
# ==============================================================================
# header synonym map -> canonical field
BUDGET_SYNONYMS = {
    "function plant": "plant",
    "function 1": "function",
    "department": "department",
    "sub department": "subDepartment",
    "position lvl": "positionLevel",
    "position id": "positionId",
    "position name": "positionName",
    "emp. code": "_empCode",      # PII -> dropped
    "emp. name": "_empName",      # PII -> masked
    "employee grade": "grade",
    "employee level": "grade",
    "regular / trainee": "employeeType",
    "regular/trainee": "employeeType",
    "vacant/occupied": "occupancy",
    "occupied/vacant": "occupancy",
    "remarks": "remarks",
    "sub department": "subDepartment",
}

def map_header(h):
    key = norm(h).lower()
    if key in BUDGET_SYNONYMS:
        return BUDGET_SYNONYMS[key]
    # contains-based fallbacks
    if "sub" in key and "department" in key:
        return "subDepartment"
    if "grade" in key or "level" in key:
        return "grade"
    if "vacant" in key or "occupied" in key:
        return "occupancy"
    if "regular" in key or "trainee" in key:
        return "employeeType"
    if key.startswith("function") and "1" in key:
        return "function"
    if "plant" in key:
        return "plant"
    if "position name" in key:
        return "positionName"
    if "position lvl" in key or key == "position level":
        return "positionLevel"
    if "emp" in key and "code" in key:
        return "_empCode"
    if "emp" in key and "name" in key:
        return "_empName"
    return None

def clean_occupancy(val, emp_name):
    """Normalise occupancy; infer from emp presence when the column is absent."""
    if not is_blank(val):
        v = str(val).strip().lower()
        if "occup" in v:
            return "Occupied"
        if "vacant" in v:
            return "Vacant"
    # fallback inference
    return "Occupied" if not is_blank(emp_name) else "Vacant"

def norm_emp_type(val):
    if is_blank(val):
        return "Regular"
    v = str(val).strip().lower()
    if "trainee" in v or "ojt" in v or "get" in v or "mt" in v:
        if "ojt" in v:
            return "OJT"
        return "Trainee"
    if "contract" in v or "bpo" in v:
        return "Contractual"
    return "Regular"

def slug(name):
    return re.sub(r"[^a-z0-9]+", "", str(name).strip().lower()) or "portfolio"

def parse_budget(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    used, ignored = [], []
    portfolio_display = {}                       # portfolio key -> display name (raw)
    pool_keys = set()                            # keys that are non-person pools

    for sn in wb.sheetnames:
        low = sn.strip().lower()
        if "summary" in low:                     # pivot/summary sheets carry no records
            ignored.append(sn)
            continue
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            ignored.append(sn)
            continue
        # A DETAIL sheet has a header row mentioning position + emp/function.
        hdr_idx = None
        for i, r in enumerate(rows[:6]):
            joined = " ".join(norm(c).lower() for c in r if c is not None)
            if "position" in joined and ("emp" in joined or "function" in joined):
                hdr_idx = i
                break
        if hdr_idx is None:
            ignored.append(sn)                   # not a record sheet
            continue

        # DATA-DRIVEN & NAME-FREE: the portfolio IS the sheet name. Variant-merging
        # and pool-folding happen later in main() via generic rules (build_resolver
        # / is_pool_name) so no names are baked into this file.
        display = canon(sn.strip())
        portfolio = slug(display)
        portfolio_display.setdefault(portfolio, display)
        is_ojt = "ojt" in low
        if is_pool_name(sn):
            pool_keys.add(portfolio)

        header = rows[hdr_idx]
        colmap = {}
        for ci, h in enumerate(header):
            field = map_header(h)
            if field and field not in colmap:
                colmap[field] = ci

        n_before = len(records)
        for r in rows[hdr_idx + 1:]:
            if not any(not is_blank(c) for c in r):
                continue
            def g(field):
                ci = colmap.get(field)
                return r[ci] if ci is not None and ci < len(r) else None
            pos_name = g("positionName")
            pos_id = g("positionId")
            if is_blank(pos_name) and is_blank(pos_id):
                continue
            emp_name = g("_empName")
            occ = clean_occupancy(g("occupancy"), emp_name)
            rec = {
                "portfolio": portfolio,
                "plant": norm(g("plant")) or NOT_AVAILABLE,
                "function": norm(g("function")) or NOT_AVAILABLE,
                "department": norm(g("department")) or NOT_AVAILABLE,
                "subDepartment": norm(g("subDepartment")) or NOT_AVAILABLE,
                "positionLevel": norm(g("positionLevel")) or NOT_AVAILABLE,
                "positionName": norm(pos_name) or NOT_AVAILABLE,
                "grade": norm(g("grade")) or NOT_AVAILABLE,
                "employeeType": "OJT" if is_ojt else norm_emp_type(g("employeeType")),
                "occupancy": occ,
                # PII masked:  emp code DROPPED entirely; emp name -> stable pseudonym
                "holder": MASK.employee(emp_name) if occ == "Occupied" else None,
                "remarks": norm(g("remarks")) or "",
            }
            records.append(rec)
        used.append(sn)
        if len(records) - n_before == 0:
            warn("budget sheet %r yielded 0 records" % sn)
    wb.close()
    return records, used, ignored, portfolio_display, pool_keys


# ==============================================================================
# 4. RECRUITMENT TRACKER — the live, record-level recruitment source (REAL data).
#    Per-role status, function, location, grade, sourcing, ageing, TAT, criticality.
#    Each row is attributed to an HRBP via the "FPR from TA/BHR" column.
# ==============================================================================
# Normalise the real, free-typed categorical values into a stable taxonomy.
STATUS_MAP = {
    "wip": "WIP", "yet to start": "Yet to Start", "tbo": "To Be Offered",
    "to be offered": "To Be Offered", "offered": "Offered", "joined": "Joined",
    "confirmation": "Confirmation", "hold": "Hold", "internal movement": "Internal Movement",
}
PIPELINE_ORDER = ["Yet to Start", "WIP", "To Be Offered", "Offered", "Joined", "Confirmation"]
PIPELINE_SIDE = ["Hold", "Internal Movement"]

def norm_status(v):
    if is_blank(v):
        return "Unknown"
    k = re.sub(r"\s+", " ", str(v)).strip().lower()
    return STATUS_MAP.get(k, str(v).strip().title())

def norm_criticality(v):
    if is_blank(v):
        return NOT_AVAILABLE
    k = str(v).strip().lower()
    if k in ("1", "high", "p1", "critical"):
        return "High"
    if k in ("2", "medium", "med", "p2"):
        return "Medium"
    if k in ("3", "low", "p3"):
        return "Low"
    if k in ("na", "n/a", "#n/a", "`", "-"):
        return NOT_AVAILABLE
    return str(v).strip().title()

def norm_postype(v):
    if is_blank(v):
        return NOT_AVAILABLE
    k = str(v).strip().lower()
    if k == "new":           return "New"
    if "replac" in k:        return "Replacement"
    if "carry" in k:         return "Carry Forwarded"
    if "adhoc" in k or "ad hoc" in k: return "Adhoc"
    return str(v).strip().title()

def attribute_hrbp(raw, portfolio_keys, resolve=None):
    """Map the 'FPR from TA/BHR' value (the HRBP's name) to a portfolio key by
    its first name. Uses the generic name resolver when supplied (so variant
    spellings match), else a plain slug. The raw name is used only for matching
    and is never stored or emitted."""
    if is_blank(raw):
        return None
    first = canon(str(raw).split()[0])
    k = resolve(first) if resolve else slug(first)
    if k is None:
        k = slug(first)
    return k if k in portfolio_keys else None

def parse_tracker(path, portfolio_keys=None, resolve=None):
    portfolio_keys = portfolio_keys or set()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = find_sheet(wb, "Recruitment Tracker ", "Recruitment Tracker")
    used = [sn] if sn else []
    ignored = [s for s in wb.sheetnames if s != sn]
    records = []
    ageing_values = []
    unattributed = 0
    bucket_mismatch = 0
    if sn:
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        header = rows[0] if rows else []
        idx = {norm(h).lower(): i for i, h in enumerate(header)}
        def col(*hints):
            # exact match first, then SHORTEST key that starts with the hint
            # (so "position" picks the title, not "position code"), then contains.
            for hint in hints:
                if hint in idx:
                    return idx[hint]
            for hint in hints:
                cands = [(len(k), i) for k, i in idx.items() if k.startswith(hint)]
                if cands:
                    return min(cands)[1]
            for hint in hints:
                for k, i in idx.items():
                    if hint in k:
                        return i
            return None
        i_bhr = col("fpr from", "bhr")
        i_appr = col("approval")
        i_year = col("budgeted year")
        i_type = col("replacement/ new", "replacement")
        i_pos = idx.get("position")                 # role TITLE (exact); NOT position code
        i_grade = col("grade")
        i_func = col("function")
        i_subfn = col("sub function")
        i_loc = col("location")
        i_src = col("sourcing through")
        i_tat = col("agreed tat for offer", "agreed tat")
        i_ageing = idx.get("ageing")
        i_bucket = col("ageing bucket")
        i_status = col("current status")
        i_crit = col("criticality / priority", "criticality")
        i_cand = col("candidate name")
        i_act = col("hiring activation")
        i_jd = col("jd finalisation")
        i_commit = col("commitment date")
        i_join = col("candidate joining")

        def date_iso(v):
            if isinstance(v, (_dt.datetime, _dt.date)):
                return v.strftime("%Y-%m-%d")
            return None

        for ri, r in enumerate(rows[1:], start=1):
            if not any(not is_blank(c) for c in r):
                continue
            def g(i):
                return r[i] if i is not None and i < len(r) else None
            ageing = num(g(i_ageing))
            if ageing is not None:
                ageing_values.append(ageing)
            tat = num(g(i_tat))
            tat = tat if (tat is not None and tat > 0) else None   # drop the -189 outlier
            # Ageing bucket: DERIVE from the numeric Ageing (days) when present so
            # the bucket always matches the days; otherwise fall back to the
            # tracker's own 'Ageing Bucket' label. Track disagreements for QC.
            src_bucket = norm(g(i_bucket))
            if ageing is not None:
                bucket = ageing_bucket(ageing)
                if src_bucket and norm_bucket_label(src_bucket) != bucket:
                    bucket_mismatch += 1
            else:
                bucket = norm_bucket_label(src_bucket) or NOT_AVAILABLE
            pf = attribute_hrbp(g(i_bhr), portfolio_keys, resolve)
            if pf is None:
                unattributed += 1
            records.append({
                "roleId": MASK.role(ri),
                "portfolio": pf,                                   # HRBP attribution
                "position": norm(g(i_pos)) or NOT_AVAILABLE,       # role title (safe)
                "function": norm(g(i_func)) or NOT_AVAILABLE,
                "subFunction": norm(g(i_subfn)) or NOT_AVAILABLE,
                "location": norm(g(i_loc)) or NOT_AVAILABLE,
                "grade": norm(g(i_grade)) or NOT_AVAILABLE,
                "positionType": norm_postype(g(i_type)),
                "budgetedYear": norm(g(i_year)) or NOT_AVAILABLE,
                "approval": norm(g(i_appr)) or NOT_AVAILABLE,
                "sourcing": norm(g(i_src)) or NOT_AVAILABLE,
                "status": norm_status(g(i_status)),
                "criticality": norm_criticality(g(i_crit)),
                "ageing": int(ageing) if ageing is not None else None,
                "ageingBucket": bucket,
                "agreedTat": int(tat) if tat is not None else None,
                "tatBreach": (ageing is not None and tat is not None and ageing > tat),
                "activationDate": date_iso(g(i_act)),
                "jdDate": date_iso(g(i_jd)),
                "commitmentDate": date_iso(g(i_commit)),
                "joiningDate": date_iso(g(i_join)),
                "candidate": MASK.candidate(g(i_cand)),            # PII masked
                # 'Reports to' (supervisor name) is intentionally NOT emitted.
            })
    wb.close()
    return records, ageing_values, used, ignored, unattributed, bucket_mismatch

def ageing_bucket(a):
    # Mutually-exclusive buckets: every role lands in exactly one. The labels do
    # NOT share endpoints (0-30, then 31-60, …) so nothing is read as counted twice.
    if a is None:
        return NOT_AVAILABLE
    a = float(a)
    if a <= 30:  return "0-30"
    if a <= 60:  return "31-60"
    if a <= 90:  return "61-90"
    if a <= 120: return "91-120"
    return "121+"

# Map the tracker's own text 'Ageing Bucket' labels (shared-endpoint style) onto the
# dashboard's mutually-exclusive labels, so fallback rows (blank numeric days) bucket
# consistently with the days-derived ones.
_SRC_BUCKET_MAP = {
    "0-30": "0-30", "30-60": "31-60", "60-90": "61-90",
    "90-120": "91-120", "120+": "121+",
}
def norm_bucket_label(s):
    s = (s or "").strip()
    return _SRC_BUCKET_MAP.get(s, s)


# ==============================================================================
# 5. RECONCILE review HRBPs <-> budget portfolios (BY NAME), then ASSEMBLE.
# ==============================================================================
def reconcile(review, portfolio_display, resolve=None):
    """Join review HRBP labels to budget portfolios by name, generically. The
    name resolver collapses spelling variants; otherwise an exact (case/space-
    insensitive) match is used. Returns an ordered list of descriptors."""
    name_to_key = {}
    for k, disp in portfolio_display.items():
        name_to_key[canon(disp).lower()] = k
        name_to_key[k] = k
    descriptors = {}
    for L in review.get("hrbpNames", []):
        k = None
        if resolve:
            rk = resolve(L)
            if rk in portfolio_display:
                k = rk
        if k is None:
            k = name_to_key.get(canon(L).lower()) or name_to_key.get(slug(L))
        if k is None:                                    # review-only HRBP
            k = slug(L)
            portfolio_display.setdefault(k, L)
            conf = "review-only"
        else:
            direct = canon(L).lower() == canon(portfolio_display.get(k, "")).lower()
            conf = "direct" if direct else "alias"
        descriptors[k] = {"key": k, "display": portfolio_display.get(k, L),
                          "reviewLabel": L, "budgetSheet": portfolio_display.get(k, NOT_AVAILABLE),
                          "confidence": conf, "verify": False}
    for k, disp in portfolio_display.items():            # budget-only portfolios
        descriptors.setdefault(k, {"key": k, "display": disp, "reviewLabel": None,
                                   "budgetSheet": disp, "confidence": "budget-only", "verify": False})
    return descriptors

CLOSED_STATUSES = ("Joined", "Confirmation", "Internal Movement")
def is_open(r):
    """Open = a role still being worked (not filled / confirmed / moved)."""
    return r["status"] not in CLOSED_STATUSES

# Ageing (and its derivatives: ageing buckets, 90+, TAT breach) is only meaningful
# for roles actively in the hiring window: WIP and To Be Offered. Every other status
# (Yet to Start, Offered, Hold, Joined, Confirmation, Internal Movement) is excluded
# from ageing calculations.
AGEING_STATUSES = ("WIP", "To Be Offered")
def is_ageing(r):
    return r["status"] in AGEING_STATUSES

def tracker_summary(recs):
    """Light per-portfolio recruitment summary from the real tracker records.
    Ageing metrics are scoped to WIP + To Be Offered only."""
    openr = [r for r in recs if is_open(r)]
    ager = [r for r in recs if is_ageing(r)]
    return {
        "total": len(recs),
        "open": len(openr),
        "wip": sum(1 for r in recs if r["status"] == "WIP"),
        "offered": sum(1 for r in recs if r["status"] == "Offered"),
        "joined": sum(1 for r in recs if r["status"] == "Joined"),
        "hold": sum(1 for r in recs if r["status"] == "Hold"),
        # ageing counts: WIP + To Be Offered only
        "ageing": len(ager),
        "ageing90plus": sum(1 for r in ager if r["ageingBucket"] in ("91-120", "121+")),
        "tatBreach": sum(1 for r in ager if r["tatBreach"]),
        # criticality (not ageing) stays scoped to all open roles
        "highCrit": sum(1 for r in openr if r["criticality"] == "High"),
    }

def assemble(review, budget_records, portfolio_display, tracker_records=None, resolve=None):
    portfolios = []
    by_pf = {}
    for r in budget_records:
        by_pf.setdefault(r["portfolio"], []).append(r)
    trk_by_pf = {}
    for r in (tracker_records or []):
        if r.get("portfolio"):
            trk_by_pf.setdefault(r["portfolio"], []).append(r)

    descriptors = reconcile(review, portfolio_display, resolve)
    # order by budget size desc, then by record count desc
    def order_key(d):
        rv = review["hrbps"].get(d["reviewLabel"], {}) if d["reviewLabel"] else {}
        return (-(rv.get("budget") or 0), -len(by_pf.get(d["key"], [])))
    ordered = sorted(descriptors.values(), key=order_key)
    review["_descriptors"] = ordered             # exposed for meta.hrbpMap

    for m in ordered:
        pkey = m["key"]
        rl = m["reviewLabel"]
        rv = review["hrbps"].get(rl, {}) if rl else {}
        recs = by_pf.get(pkey, [])

        budget = rv.get("budget")
        active = rv.get("active")
        vacancy = (budget - active) if (budget is not None and active is not None) else None
        vacancyPct = round(vacancy / budget * 100, 1) if (vacancy is not None and budget) else None

        rec = rv.get("recruitment", {})
        def rsum(stage):
            d = rec.get(stage)
            return d["total"] if d else 0
        open_pipeline = rsum("wip") + rsum("toBeOffered") + rsum("offered")
        offered = rsum("offered")
        joined = rsum("joined")
        offer_to_join = round(joined / offered * 100, 1) if offered else None
        total_funnel = open_pipeline + joined + rsum("offerDeclined")
        closure = round(joined / total_funnel * 100, 1) if total_funnel else None

        aging = rv.get("aging", {})
        wip = aging.get("wip", 0)
        b90 = aging.get("b90", 0)
        tat_breach_pct = round(b90 / wip * 100, 1) if wip else None

        pms = rv.get("pms", {})
        goal = pms.get("goalSetting")
        pms_pending = round(100 - goal, 1) if goal is not None else None

        attrition = rv.get("attrition")

        # engagement: lowest dept score vs benchmark
        eng = [e for e in review["engagement"]
               if rl and e["hrbp"].lower() == rl.lower() and e["score"] is not None]
        lowest = min(eng, key=lambda e: e["score"]) if eng else None

        trk = tracker_summary(trk_by_pf.get(pkey, []))

        portfolios.append({
            "key": pkey,
            "display": m["display"],
            "reviewLabel": rl,
            "budgetSheet": m["budgetSheet"],
            "confidence": m["confidence"],
            "verify": m["verify"],
            "tracker": trk,
            # headcount
            "budget": budget, "active": active,
            "vacancy": vacancy, "vacancyPct": vacancyPct,
            "joiningsJune": rv.get("joiningsJune"), "exitsJune": rv.get("exitsJune"),
            "joiningsYTD": rv.get("joiningsYTD"), "exitsYTD": rv.get("exitsYTD"),
            "netMovementYTD": ((rv.get("joiningsYTD") or 0) - (rv.get("exitsYTD") or 0)),
            "future": rv.get("future"), "onHold": rv.get("onHold"),
            "delimit": rv.get("delimit"),
            "attrition": attrition,
            "attritionInsight": rv.get("attritionInsight", NOT_AVAILABLE),
            # recruitment funnel
            "recruitment": rec,
            "openPipeline": open_pipeline, "offered": offered, "joined": joined,
            "offerToJoin": offer_to_join, "closureRate": closure,
            # ageing
            "aging": aging, "tatBreachPct": tat_breach_pct,
            "criticalCases": rv.get("criticalCases", []),
            # pms
            "pms": pms, "pmsPending": pms_pending,
            # training / org / engagement
            "training": rv.get("training"),
            "orgChart": rv.get("orgChart", NOT_AVAILABLE),
            "engagementLowest": (lowest["department"] if lowest else NOT_AVAILABLE),
            "engagementLowestScore": (lowest["score"] if lowest else None),
            # budget-record rollups
            "budgetRecordCount": len(recs),
            "occupiedRecords": sum(1 for r in recs if r["occupancy"] == "Occupied"),
            "vacantRecords": sum(1 for r in recs if r["occupancy"] == "Vacant"),
        })

    # Portfolio Risk Index (workload/risk indicator — NOT performance).
    # weights: Vacancy 25 · Ageing 25 · Attrition 20 · Hiring load 15 · PMS pending 10 · Engagement 5
    def safe(v, d=0): return v if v is not None else d
    max_vac = max([safe(p["vacancyPct"]) for p in portfolios] + [1])
    max_age = max([safe(p["aging"].get("b90")) for p in portfolios] + [1])
    max_attr = max([safe(p["attrition"]) for p in portfolios] + [1])
    max_load = max([safe(p["openPipeline"]) for p in portfolios] + [1])
    max_pms = max([safe(p["pmsPending"]) for p in portfolios] + [1])
    bench = review.get("benchmark") or 6.19
    for p in portfolios:
        eng_gap = max(0, bench - safe(p["engagementLowestScore"], bench))
        idx = (
            25 * safe(p["vacancyPct"]) / max_vac +
            25 * safe(p["aging"].get("b90")) / max_age +
            20 * safe(p["attrition"]) / max_attr +
            15 * safe(p["openPipeline"]) / max_load +
            10 * safe(p["pmsPending"]) / max_pms +
            5 * (eng_gap / max(bench, 1))
        )
        p["riskIndex"] = round(idx, 1)
        p["riskBand"] = ("High" if idx >= 60 else "Moderate" if idx >= 35 else "Low")
    return portfolios


# ==============================================================================
# 6. KPI roll-up (All HRBPs) + actions + data quality
# ==============================================================================
def rollup_kpis(portfolios, review):
    def s(field):
        return sum((p[field] or 0) for p in portfolios if p.get(field) is not None)
    budget = s("budget"); active = s("active")
    vacancy = budget - active
    joinYTD = s("joiningsYTD"); exitYTD = s("exitsYTD")
    openpipe = s("openPipeline")
    # weighted-average attrition (by active)
    tot_active = sum((p["active"] or 0) for p in portfolios) or 1
    attr = round(sum((p["attrition"] or 0) * (p["active"] or 0) for p in portfolios) / tot_active, 1)
    return {
        "budget": budget, "active": active,
        "vacancy": vacancy,
        "vacancyPct": round(vacancy / budget * 100, 1) if budget else None,
        "joiningsYTD": joinYTD, "exitsYTD": exitYTD,
        "netMovementYTD": joinYTD - exitYTD,
        "attrition": attr,
        "openPipeline": openpipe,
        "joined": s("joined"), "offered": s("offered"),
        "benchmark": review.get("benchmark") or 6.19,
    }

def build_actions(portfolios):
    """Derive 'This Month's Priorities' from the signals."""
    actions = []
    aid = 0
    def add(pf, theme, issue, evidence, impact, rec, owner, priority):
        nonlocal aid
        aid += 1
        actions.append({
            "id": "A%03d" % aid, "hrbp": pf["display"], "theme": theme,
            "issue": issue, "evidence": evidence, "impact": impact,
            "recommendation": rec, "owner": owner, "due": "This month",
            "priority": priority, "status": "Not Started",
        })
    for p in portfolios:
        if p["vacancyPct"] is not None and p["vacancyPct"] > 20:
            add(p, "Vacancy", "Vacancy pressure above 20%%",
                "Vacancy %.1f%% (%s of %s positions)" % (p["vacancyPct"], p["vacancy"], p["budget"]),
                "Capacity & delivery risk", "Prioritise sourcing for open roles; review hold list", "HRBP / TA", "P1")
        b90 = p["aging"].get("b90", 0)
        if b90 and b90 >= 10:
            add(p, "Ageing", "Roles ageing beyond 90 days",
                "%d roles in 90+ bucket" % b90, "TAT breach / business escalation",
                "Escalate 90+ roles; revisit sourcing channel", "TA", "P1")
        if p["attrition"] is not None and p["attrition"] > 5:
            add(p, "Attrition", "Attrition above 5%%",
                "Attrition %.1f%%" % p["attrition"], "Talent loss",
                "Run retention check-ins; act on exit themes", "HRBP", "P1")
        goal = (p["pms"] or {}).get("goalSetting")
        if goal is not None and goal < 85:
            add(p, "PMS", "PMS goal-setting below 85%%",
                "Goal setting %.1f%%" % goal, "Performance cycle readiness",
                "Drive goal-setting completion", "HRBP", "P2")
        if p["criticalCases"]:
            add(p, "Business dependency", "Critical roles flagged",
                "; ".join(p["criticalCases"][:3]), "Key role gaps",
                "Track top-3 critical roles to closure", "HRBP / Business", "P1")
        trk = p.get("tracker") or {}
        if trk.get("tatBreach"):
            add(p, "TAT", "Roles open beyond agreed TAT",
                "%d roles past their agreed TAT" % trk["tatBreach"], "Hiring SLA breach",
                "Escalate TAT-breached roles to TA / business", "TA", "P1")
        if trk.get("highCrit"):
            add(p, "Criticality", "High-criticality roles open",
                "%d high-criticality roles in pipeline" % trk["highCrit"], "Business-critical gaps",
                "Prioritise high-criticality roles this month", "HRBP / TA", "P1")
    return actions

def data_quality(tracker_records, budget_records, portfolios, ageing_values, unattributed=0, bucket_mismatch=0):
    issues = []
    def add(kind, count, detail, severity):
        issues.append({"type": kind, "count": count, "detail": detail, "severity": severity})

    n_tr = len(tracker_records)
    # Ageing only applies to WIP + To Be Offered — measure the gap on those rows.
    ageing_scope = [r for r in tracker_records if r["status"] in AGEING_STATUSES]
    miss_ageing = sum(1 for r in ageing_scope if r["ageing"] is None)
    miss_bucket = sum(1 for r in ageing_scope if r["ageingBucket"] == NOT_AVAILABLE)
    miss_crit = sum(1 for r in tracker_records if r["criticality"] == NOT_AVAILABLE)
    miss_status = sum(1 for r in tracker_records if r["status"] == "Unknown")
    non_appr = sum(1 for r in tracker_records if "non" in (r["approval"] or "").lower())
    miss_tat = sum(1 for r in tracker_records if r["agreedTat"] is None)
    joined_no_date = sum(1 for r in tracker_records if r["status"] == "Joined" and not r["joiningDate"])

    add("Missing ageing (tracker)", miss_ageing,
        "WIP / To-Be-Offered roles with no numeric ageing — ageing-day metrics exclude these.", "medium")
    add("Missing ageing bucket", miss_bucket, "WIP / To-Be-Offered rows with no ageing bucket.", "low")
    add("Missing criticality", miss_crit, "Roles with no Criticality/Priority set.", "medium")
    add("Unknown status", miss_status, "Rows whose Current Status is blank/unmapped.", "medium")
    add("Non-approved roles", non_appr, "Roles flagged Non-Approved still in the pipeline.", "low")
    add("Joined without joining date", joined_no_date, "Status = Joined but no joining date.", "low")
    add("Missing agreed TAT", miss_tat, "Rows with no (or invalid) agreed TAT — TAT-breach excludes these.", "low")
    if unattributed:
        add("Unattributed tracker rows", unattributed,
            "Rows whose 'FPR from TA/BHR' did not match a portfolio (shown under All only).", "medium")
    if bucket_mismatch:
        add("Ageing bucket vs days mismatch", bucket_mismatch,
            "Rows where the tracker's 'Ageing Bucket' label disagreed with its numeric "
            "'Ageing' (days). The dashboard uses the days-derived bucket for these.", "medium")
    miss_grade = sum(1 for r in budget_records if r["grade"] == NOT_AVAILABLE)
    add("Missing grade (budget)", miss_grade, "Budget rows without a usable grade.", "low")

    total = n_tr if n_tr else 1
    # completeness over the key actionable fields
    completeness = round(100 * (1 - (miss_ageing + miss_crit + miss_status) / (3 * total)), 1) if n_tr else 100.0
    return {
        "completeness": completeness,
        "trackerRows": n_tr,
        "budgetRows": len(budget_records),
        "ageingMin": min(ageing_values) if ageing_values else None,
        "ageingMax": max(ageing_values) if ageing_values else None,
        "unattributed": unattributed,
        "issues": issues,
        "actionRequired": sum(1 for i in issues if i["severity"] in ("high", "medium")),
    }


# ==============================================================================
# 7. OUTPUT GUARD — fail loudly if PII leaks into data.js
# ==============================================================================
# Enum / common-text words that may coincide with a (masked) first name but are
# NOT a leak when they appear in the output (statuses, programmes, etc.).
GUARD_SAFE_WORDS = {
    "hold", "joined", "offered", "other", "lead", "leap", "seed", "succeed",
    "new", "none", "review", "pending", "confirmation", "internal", "movement",
    "medium", "approved", "shift", "sales", "legal", "audit", "stores", "gift",
}
def output_guard(js_text, source_names):
    problems = []
    low = js_text.lower()
    for name in source_names:
        nm = re.sub(r"\s+", " ", str(name)).strip().lower()
        if len(nm) < 5 or "demo" in nm or "dummy" in nm:
            continue
        if nm in GUARD_SAFE_WORDS:
            continue
        # whole-word/standalone match only (avoids "dev" inside "Development")
        if re.search(r"\b" + re.escape(nm) + r"\b", low):
            problems.append("source name leaked: %r" % name)
    # emp-code-shaped 6-8 digit runs inside quoted values (codes are dropped)
    for m in re.finditer(r'"[^"]*\b(\d{6,8})\b[^"]*"', js_text):
        problems.append("possible emp-code-shaped value: %s" % m.group(0)[:60])
    return problems


# ==============================================================================
# MAIN
# ==============================================================================
def resolve(src, fname):
    for cand in (os.path.join(src, fname), os.path.join(src, "data", fname)):
        if os.path.exists(cand):
            return cand
    return None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=".", help="folder containing the 3 xlsx (or its ./data)")
    ap.add_argument("--out", default=os.path.join("js", "data.js"))
    args = ap.parse_args()

    print("=" * 78)
    print("HRBP Command Centre — data build")
    print("=" * 78)

    p_review = resolve(args.src, FILE_REVIEW)
    p_budget = resolve(args.src, FILE_BUDGET)
    p_tracker = resolve(args.src, FILE_TRACKER)
    for label, p in [("Monthly Review", p_review), ("Budget", p_budget), ("Tracker", p_tracker)]:
        if not p:
            sys.exit("ERROR: could not find %s workbook near %r" % (label, args.src))
        print("  read:", p)

    print("\n[1/5] Parsing Monthly Review (PRIMARY)...")
    review = parse_review(p_review)
    print("      HRBPs:", list(review["hrbps"].keys()),
          "| engagement depts:", len(review["engagement"]),
          "| initiatives:", len(review["initiatives"]))

    print("[2/5] Parsing Budget workbook (record-level)...")
    budget_records, b_used, b_ignored, portfolio_display, pool_keys = parse_budget(p_budget)
    print("      detail sheets used:", b_used)
    print("      records:", len(budget_records))

    # --- NAME-FREE canonicalisation: merge spelling variants + fold pool sheets ---
    fold_cfg, merge_cfg = load_local_overrides(args.src)
    person_keys = [k for k in portfolio_display if k not in pool_keys]
    resolver_names = [portfolio_display[k] for k in person_keys] + list(review.get("hrbpNames", []))
    name_resolve, res_disp = build_resolver(resolver_names)

    # old budget key -> canonical person key; canonical display = shortest form
    # (res_disp already holds the shortest form across budget + review names)
    remap, canon_display = {}, {}
    for k in person_keys:
        disp = portfolio_display[k]
        nk = (slug(merge_cfg[disp.lower()]) if disp.lower() in merge_cfg else None) or name_resolve(disp) or k
        remap[k] = nk
        canon_display[nk] = res_disp.get(nk, disp)

    # pool sheets fold into the largest person portfolio (by budget record count),
    # unless an explicit override pins them elsewhere
    rec_count = collections.Counter(r["portfolio"] for r in budget_records)
    def folded_count(nk):
        return sum(c for k, c in rec_count.items() if remap.get(k) == nk)
    ranked = sorted(canon_display, key=lambda nk: -folded_count(nk))
    fold_target = ranked[0] if ranked else None
    for k in pool_keys:
        disp = portfolio_display[k]
        override = fold_cfg.get(disp.lower())
        remap[k] = (name_resolve(override) or slug(override)) if override else fold_target
        if remap[k]:
            print("      folded pool sheet %r -> %r" % (disp, canon_display.get(remap[k], remap[k])))

    for r in budget_records:                              # apply the remap
        r["portfolio"] = remap.get(r["portfolio"], r["portfolio"])
    portfolio_display = canon_display
    print("      portfolios discovered:", list(portfolio_display.values()))

    print("[3/5] Parsing Recruitment Tracker (record-level, real)...")
    portfolio_keys = set(portfolio_display.keys())
    tracker_records, ageing_values, t_used, t_ignored, t_unattr, t_bktmm = parse_tracker(p_tracker, portfolio_keys, name_resolve)
    attributed = sum(1 for r in tracker_records if r.get("portfolio"))
    print("      tracker rows:", len(tracker_records),
          "| attributed to HRBP:", attributed, "| unattributed:", t_unattr,
          "| numeric ageing:", len(ageing_values))

    print("[4/5] Assembling portfolios + KPIs...")
    portfolios = assemble(review, budget_records, portfolio_display, tracker_records, name_resolve)
    kpis = rollup_kpis(portfolios, review)
    actions = build_actions(portfolios)
    dq = data_quality(tracker_records, budget_records, portfolios, ageing_values, t_unattr, t_bktmm)

    hrbp_map_view = []
    for m in review.get("_descriptors", []):
        hrbp_map_view.append({
            "display": m["display"], "reviewLabel": m["reviewLabel"] or NOT_AVAILABLE,
            "budgetSheet": m["budgetSheet"], "confidence": m["confidence"],
            "verify": m["verify"],
        })

    data = {
        "meta": {
            "title": "HRBP Workforce & Recruitment Command Centre",
            "subtitle": "Monthly Review | Budget • Hiring • Attrition • Capability • Engagement",
            "generatedAt": _dt.date.today().isoformat(),
            "tatAssumptionDays": DEFAULT_TAT_DAYS,
            "benchmark": kpis["benchmark"],
            "hrbpMap": hrbp_map_view,
            "sources": {
                "review": FILE_REVIEW, "budget": FILE_BUDGET, "tracker": FILE_TRACKER,
            },
            "budgetSheetsUsed": b_used,
            "ignoredSheets": sorted(set(b_ignored + t_ignored)),
            "warnings": WARNINGS,
            "privacyNote": "All names, employee codes, candidate names and supervisor "
                           "references are masked at bake time. No PII is shipped.",
        },
        "kpis": kpis,
        "portfolios": portfolios,
        "budgetRecords": budget_records,
        "recruitmentRecords": tracker_records,
        "engagement": review["engagement"],
        "speakUpMilestones": review.get("speakUpMilestones", []),
        "initiatives": review["initiatives"],
        "actions": actions,
        "dataQuality": dq,
    }

    print("[5/5] Emitting %s + running output guard..." % args.out)
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    header = ("// generated %s — do not edit by hand\n"
              "// HRBP Workforce & Recruitment Command Centre — baked, anonymised data.\n"
              % _dt.date.today().isoformat())
    js_text = header + "window.DASHBOARD_DATA = " + payload + ";\n"

    problems = output_guard(payload, MASK.source_name_values())
    if problems:
        print("\nOUTPUT GUARD FAILED — PII leak suspected:")
        for pr in problems[:20]:
            print("   -", pr)
        sys.exit("Build aborted: refusing to emit data.js with possible PII.")

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(js_text)

    # ---- console summary ----
    print("\n" + "-" * 78)
    print("BUILD SUMMARY")
    print("-" * 78)
    print("  Files read : %s, %s, %s" % (FILE_REVIEW, FILE_BUDGET, FILE_TRACKER))
    print("  Budget detail sheets used :", ", ".join(b_used))
    print("  Ignored / pivot sheets    :", ", ".join(sorted(set(b_ignored + t_ignored))))
    print("  Budget records   :", len(budget_records))
    print("  Tracker records  :", len(tracker_records), "(record-level, HRBP-attributed)")
    print("  Engagement depts :", len(review["engagement"]))
    print("  Initiatives      :", len(review["initiatives"]))
    print("  Derived actions  :", len(actions))
    print("  HRBP join (portfolio <- review label) [match type]:")
    for v in hrbp_map_view:
        print("     %-12s <- review:%-10s [%s]"
              % (v["display"], v["reviewLabel"], v["confidence"]))
    print("  Data-quality completeness :", dq["completeness"], "%")
    if WARNINGS:
        print("  Warnings:")
        for w in WARNINGS:
            print("     -", w)
    print("  Output guard : PASSED (no PII leaked)")
    print("  Wrote        :", args.out, "(%d KB)" % (len(js_text) // 1024))
    print("-" * 78)
    print("Done.")

if __name__ == "__main__":
    main()
